import openpyxl
import random
import os
import sys

file_name = ''
file = os.path.isfile(file_name)

class OpenExcel():
    def __init__(self):
        workbook = openpyxl.load_workbook(file_name)
        worksheet = workbook['Munka1']

        self.Datas = []
        data = {}
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            data = {}
            data['Id'] = row[1]
            data['Name'] = row[2]
            data['Position'] = row[3]
            data['Supervisor'] = row[4]
            data['Start'] = str(row[5])
            data['Benefits'] = row[6]
            self.Datas.append(data)
    def Out(self):
        for data in self.Datas:
            print(f"Id : {data['Id']}, \t Name : {data['Name']}, \t Position : {data['Position']}, \t Supervisor : {data['Supervisor']}, \t Start : {data['Start']}, \t {data['Benefits']}")

class CreateExcel():
    def __init__(self):
        self.book = openpyxl.Workbook()
        sheet = self.book.active
        sheet['B1'] = 'Id'
        sheet['C1'] = 'Name'
        sheet['D1'] = 'Position'
        sheet['E1'] = 'Supervisor'
        sheet['F1'] = 'Start'
        sheet['G1'] = 'Benefits'
        self.position = {1: 'Owner', 2: 'DepartmentLeader', 3: 'Teamleader', 4: 'Developer'}
        self.position_list = []
        self.benefits = {1: 'Car', 2: 'Phone', 3: 'Fuel card', 4: 'Gift card'}
    def owner(self):
        self.position_line_data = {}
        self.position_line_data['Id'] = 2
        self.position_line_data['Name'] = 'Voros Bence'
        self.position_line_data['Position'] = self.position[1]
        self.position_line_data['Supervisor'] = None
        self.position_line_data['Start'] = '2010.10.11'
        self.position_line_data['Benefits'] = self.benefits[random.randint(1, 4)]
        self.position_list.append(self.position_line_data)
    def generate(self,max_num):
        employees = {}
        employees['Leader'] = 1
        employees['DepartmentLeader'] = round((max_num/100) * 20)
        employees['TeamLeader'] = round((max_num / 100) * 30)
        employees['Developer'] = max_num-(employees['Leader']+employees['DepartmentLeader']+employees['TeamLeader'])
        self.owner()
        self.departmentleader(employees['DepartmentLeader'])
        self.teamLeader(self.dp_number_id,employees['TeamLeader'])
        self.developer(self.tl_number_id,employees['Developer'])
        self.addToExcel()
    def departmentleader(self,num):
        self.dp_number_id = 3+num
        for id in range(3,self.dp_number_id+1):
            self.position_line_data = {}
            self.position_line_data['Id'] = id
            self.position_line_data['Name'] = self.nameGenerator()
            self.position_line_data['Position'] = self.position[2]
            self.position_line_data['Supervisor'] = self.position_list[0]['Id']
            self.position_line_data['Start'] = self.dateGenerator()
            self.position_line_data['Benefits'] = self.benefits[random.randint(1, 4)]
            self.position_list.append(self.position_line_data)
    def teamLeader(self,next_num,num_id):
        self.tl_number_id = next_num + num_id
        for id in range(next_num+1,self.tl_number_id+1):
            self.position_line_data = {}
            self.position_line_data['Id'] = id
            self.position_line_data['Name'] = self.nameGenerator()
            self.position_line_data['Position'] = self.position[3]
            self.position_line_data['Supervisor'] = self.position_list[random.randint(1,self.dp_number_id-2)]['Id']
            self.position_line_data['Start'] = self.dateGenerator()
            self.position_line_data['Benefits'] = self.benefits[random.randint(1, 4)]
            self.position_list.append(self.position_line_data)
    def developer(self,next_num,num_id):
        self.dev_number_id = next_num + num_id
        for id in range(next_num+1,self.dev_number_id):
            self.position_line_data = {}
            self.position_line_data['Id'] = id
            self.position_line_data['Name'] = self.nameGenerator()
            self.position_line_data['Position'] = self.position[4]
            self.position_line_data['Supervisor'] = self.position_list[random.randint(self.dp_number_id-1,self.tl_number_id-2)]['Id']
            self.position_line_data['Start'] = self.dateGenerator()
            self.position_line_data['Benefits'] = self.benefits[random.randint(1, 4)]
            self.position_list.append(self.position_line_data)
    def dateGenerator(self):
        year = random.randint(1980,2023)
        month = random.randint(1,12)
        if month < 10:
            month = '0'+str(month)
            int(month)
        if month == 2:
            day = random.randint(1,28)
            if day < 10:
                day = '0'+str(day)
                int(day)
        elif month == 2 and year % 4 == 0:
            day = random.randint(1,29)
            if day < 10:
                day = '0'+str(day)
                int(day)
        elif month == 1 or month == 3 or month == 5 or month == 7 or month == 8 or month == 10 or month == 12:
            day = random.randint(1,31)
            if day < 10:
                day = '0'+str(day)
                int(day)
        else:
            day = random.randint(1,30)
            if day < 10:
                day = '0'+str(day)
                int(day)
        date = f'{year}.{month}.{day}'
        return date
    def nameGenerator(self):
        names = {1: 'Oliver', 2: 'Jack', 3: 'Harry', 4: 'Jacob', 5: 'Charlie', 6: 'Thomas', 7: 'George', 8: 'Oscar',9:'Margaret', 10:'Elizabeth'}
        return names[random.randint(1,10)]

    def addToExcel(self):
        sheet = self.book.active
        sheet['B2'] = f'000{self.position_list[0]["Id"]-1}'
        sheet['C2'] = self.position_list[0]['Name']
        sheet['D2'] = self.position_list[0]['Position']
        sheet['E2'] = self.position_list[0]['Supervisor']
        sheet['F2'] = self.position_list[0]['Start']
        sheet['G2'] = self.position_list[0]['Benefits']
        for data in self.position_list:
            if len(str(int(data['Id'])-1)) == 1:
                pos_id = f'000{data["Id"]-1}'
            elif len(str(int(data['Id'])-1)) == 2:
                pos_id = f'00{data["Id"]-1}'
            elif len(str(int(data['Id'])-1)) == 3:
                pos_id = f'0{data["Id"]-1}'
            elif len(str(int(data['Id'])-1)) == 4:
                pos_id = f'{data["Id"]-1}'

            if data['Supervisor'] == None:
                continue
            elif len(str(int(data['Supervisor'])-1)) == 1:
                Supervisor_id = f'000{data["Supervisor"]-1}'
            elif len(str(int(data['Supervisor'])-1)) == 2:
                Supervisor_id = f'00{data["Supervisor"] - 1}'
            elif len(str(int(data['Supervisor'])-1)) == 3:
                Supervisor_id = f'0{data["Supervisor"] - 1}'
            elif len(str(int(data['Supervisor'])-1)) == 4:
                Supervisor_id = f'{data["Supervisor"] - 1}'


            sheet['B'+str(data["Id"])] = pos_id
            sheet['C'+str(data["Id"])] = data['Name']
            sheet['D'+str(data["Id"])] = data['Position']
            sheet['E'+str(data["Id"])] = Supervisor_id
            sheet['F'+str(data["Id"])] = data['Start']
            sheet['G'+str(data["Id"])] = data['Benefits']
    def create(self,name):
        self.book.save(name)
if file:
    openFile = OpenExcel()
    openFile.Out()
else:
    createFile = CreateExcel()
    createFile.generate(50)
    createFile.create('ilovepython.xlsx')
